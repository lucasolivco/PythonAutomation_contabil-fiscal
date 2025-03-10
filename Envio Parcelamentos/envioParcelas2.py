import os
import zipfile
import shutil
from datetime import datetime
from pywinauto import Application, timings, findwindows
import time
import win32gui
import win32con
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook

timings.Timings.window_find_timeout = 5
timings.Timings.app_start_timeout = 10
timings.Timings.exists_timeout = 1
timings.Timings.after_click_wait = 0.1
timings.Timings.after_editsetedittext_wait = 0.1

# variáveis globais:
total_pdfs = 0
pdfs_processados = 0

# Aviso de que o processo foi concluído
def mostrar_aviso():
    root = tk.Tk()
    root.withdraw()  # Esconde a janela principal
    messagebox.showinfo("Aviso", "O processamento de todas as pastas foi concluído!")
    root.destroy()

# Cria log para relatório
def registrar_log(empresa, numero, pdf_path):
    log_file = "automacao_log.txt"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(f"{timestamp} - Empresa: {empresa}, Número: {numero}, PDF: {pdf_path}\n\n")

# Função para carregar os dados da planilha Excel
def carregar_dados_excel(caminho_excel):
    wb = load_workbook(caminho_excel)
    ws = wb.active
    empresa_para_numero = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            empresa_para_numero[row[0]] = str(row[1])
    return empresa_para_numero

# Carregar dados da planilha Excel
caminho_excel = r'C:\Users\VMContabil\Documents\Python automate\Envio Parcelamentos\Associação empresas-número.xlsx'  # Substitua pelo caminho real da sua planilha
empresa_para_numero = carregar_dados_excel(caminho_excel)

def encontrar_zip_mais_recente(diretorio):
    arquivos_zip = [f for f in os.listdir(diretorio) if f.endswith('.zip')]
    if not arquivos_zip:
        return None
    return max(arquivos_zip, key=lambda f: os.path.getmtime(os.path.join(diretorio, f)))
def extrair_zip(arquivo_zip, diretorio_destino):
    try:
        with zipfile.ZipFile(arquivo_zip, 'r') as zip_ref:
            zip_ref.extractall(diretorio_destino)
        print(f"Arquivo ZIP extraído com sucesso: {arquivo_zip}")
    except zipfile.BadZipFile:
        print(f"Erro: O arquivo {arquivo_zip} não é um arquivo ZIP válido.")
    except Exception as e:
        print(f"Erro ao extrair o arquivo ZIP: {e}")

def reorganizar_pastas(diretorio_raiz):
    

    for pasta_empresa in os.listdir(diretorio_raiz):
        caminho_empresa = os.path.join(diretorio_raiz, pasta_empresa)
    
        contador_docs = 1
        for primeira_subpasta in os.listdir(caminho_empresa):
            caminho_primeira_subpasta = os.path.join(caminho_empresa, primeira_subpasta)
            if os.path.isdir(caminho_primeira_subpasta):
                for raiz, diretorios, arquivos in os.walk(caminho_primeira_subpasta):
                    for arquivo in arquivos:
                        if arquivo.lower().endswith('.pdf'):
                            caminho_antigo = os.path.join(raiz, arquivo)
                            nome_base, extensao = os.path.splitext(arquivo)
                            novo_nome = f"{nome_base}_{contador_docs:02d}{extensao}"
                            caminho_novo = os.path.join(caminho_empresa, novo_nome)

                            # Renomeia e move o arquivo
                            shutil.move(caminho_antigo, caminho_novo)
                            contador_docs += 1

            # Remove subpastas vazias
            for raiz, diretorios, arquivos in os.walk(caminho_empresa, topdown=False):
                for diretorio in diretorios:
                    caminho_diretorio = os.path.join(raiz, diretorio)
                    if not os.listdir(caminho_diretorio):
                        os.rmdir(caminho_diretorio)



# Diretório onde o script está localizado
diretorio_script = os.path.dirname(os.path.abspath(__file__))

# Encontrar o arquivo ZIP mais recente
arquivo_zip = encontrar_zip_mais_recente(diretorio_script)

def find_dominio_window():
    try:
        windows = findwindows.find_windows(title_re=".*Domínio Escrita Fiscal.*")
        if windows:
            return windows[0]
        else:
            print("Nenhuma janela do Domínio Escrita Fiscal encontrada.")
            return None
    except Exception as e:
        print(f"Erro ao procurar a janela do Domínio Escrita Fiscal: {str(e)}")
        return None

def wait_for_atencao_window(timeout=10, retry_interval=1):
    start_time = time.time()
    while time.time() - start_time < timeout:
        try:
            windows = findwindows.find_windows(title="Atenção")
            if windows:
                return windows[0]
        except:
            pass
        time.sleep(retry_interval)
    return None

def interact_with_element(element, action, value=None, max_attempts=5):
    for attempt in range(max_attempts):
        try:
            if action == "set_focus":
                element.set_focus()
            elif action == "set_edit_text":
                element.set_edit_text(value)
            elif action == "click":
                element.click()
            return True
        except Exception as e:
            print(f"Tentativa {attempt + 1} falhou: {str(e)}")
            time.sleep(1)
    print(f"Falha após {max_attempts} tentativas.")
    return False

def interact_with_dominio_escrita_fiscal(pdf_path, numero_empresa, empresa):
    global pdfs_processados
    try:
        timings.Timings.window_find_timeout = 20

        handle = find_dominio_window()
        if not handle:
            print("Não foi possível encontrar a janela do Domínio Escrita Fiscal.")
            return

        if win32gui.IsIconic(handle):
            win32gui.ShowWindow(handle, win32con.SW_RESTORE)
            time.sleep(1)

        app = Application(backend="uia").connect(handle=handle)
        main_window = app.window(handle=handle)

        pub_window = main_window.child_window(title="Publicação de Documentos Externos")

        if not pub_window.exists():
            print("Janela 'Publicação de Documentos Externos' não encontrada.")
            return

        onvio_processos = pub_window.child_window(class_name="Edit", auto_id="1013")

        if onvio_processos.exists():
            if interact_with_element(onvio_processos, "set_focus") and \
               interact_with_element(onvio_processos, "set_edit_text", pdf_path):
                print(f"Campo Onvio Processos preenchido com: {pdf_path}")
            else:
                print("Falha ao interagir com o campo Onvio Processos.")
                return
        else:
            print("Campo 'Onvio Processos' não encontrado.")
            return

        campo_numero = pub_window.child_window(class_name="PBEDIT190", auto_id="1001")

        if campo_numero.exists():
            if interact_with_element(campo_numero, "set_focus") and \
               interact_with_element(campo_numero, "set_edit_text", numero_empresa):
                print(f"Campo PBEDIT190 preenchido com o número: {numero_empresa}")
            else:
                print("Falha ao interagir com o campo PBEDIT190.")
                return
        else:
            print("Campo PBEDIT190 com AutomationId 1001 não encontrado.")

        # botao = pub_window.child_window(class_name="Button", auto_id="1003")

        # if botao.exists():
        #     if interact_with_element(botao, "click"):
        #         print("Botão com AutomationId 1003 clicado com sucesso!")

        #         # Esperar pela janela de Atenção
        #         atencao_handle = wait_for_atencao_window()
        #         if atencao_handle:
        #             print("Janela de Atenção encontrada.")
        #             atencao_app = Application(backend="uia").connect(handle=atencao_handle)
        #             atencao_window = atencao_app.window(handle=atencao_handle)
        #             atencao_window.set_focus()
        #             ok_button = atencao_window.child_window(title="OK")
        #             if ok_button.exists():
        #                 ok_button.click()
        #                 print("Botão OK clicado na janela de Atenção.")
        #                 # Registrar no log após sucesso
        #                 registrar_log(empresa, numero_empresa, pdf_path)
        #         else:
        #             print("Janela de Atenção não apareceu após 10 segundos.")
        #     else:
        #         print("Falha ao clicar no botão com AutomationId 1003.")
        # else:
        #     print("Botão com AutomationId 1003 não encontrado.")

        pdfs_processados += 1
        print(f"PDFs processados: {pdfs_processados}/{total_pdfs}")

        # Se todos os PDFs foram processados, mostre o aviso
        if pdfs_processados == total_pdfs:
            mostrar_aviso()

    except Exception as e:
        print(f"Erro ao interagir com Domínio Escrita Fiscal: {str(e)}")

# Diretório onde o script está localizado
diretorio_script = os.path.dirname(os.path.abspath(__file__))

# Encontrar o arquivo ZIP mais recente
arquivo_zip = encontrar_zip_mais_recente(diretorio_script)

if arquivo_zip:
    arquivo_zip_path = os.path.join(diretorio_script, arquivo_zip)
    diretorio_destino = os.path.join(diretorio_script, "extraido_" + datetime.now().strftime("%Y%m%d_%H%M%S"))

    print(f"Arquivo ZIP encontrado: {arquivo_zip}")
    print(f"Extraindo para: {diretorio_destino}")

    extrair_zip(arquivo_zip_path, diretorio_destino)
    reorganizar_pastas(diretorio_destino)
    print("Reorganização concluída!")

    # Contar o número total de PDFs
    total_pdfs = sum(len([f for f in files if f.lower().endswith('.pdf')])
                     for _, _, files in os.walk(diretorio_destino))
    print(f"Total de PDFs encontrados: {total_pdfs}")

    # Interagir com Domínio Escrita Fiscal para cada PDF
    for empresa in os.listdir(diretorio_destino):
        caminho_empresa = os.path.join(diretorio_destino, empresa)
        if os.path.isdir(caminho_empresa):
            numero_empresa = empresa_para_numero.get(empresa, "000")

            for arquivo in os.listdir(caminho_empresa):
                if arquivo.lower().endswith('.pdf'):
                    pdf_path = os.path.join(caminho_empresa, arquivo)
                    interact_with_dominio_escrita_fiscal(pdf_path, numero_empresa, empresa)

    # Verificação final caso algum PDF não tenha sido processado
    if pdfs_processados < total_pdfs:
        print(f"Aviso: Nem todos os PDFs foram processados. Processados: {pdfs_processados}/{total_pdfs}")

else:
    print("Nenhum arquivo ZIP encontrado no diretório do script.")
