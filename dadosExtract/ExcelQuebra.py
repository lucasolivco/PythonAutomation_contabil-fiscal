import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def apply_wrap_text(input_folder, output_folder):
    # Certifique-se de que a pasta de saída existe
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Percorra todos os arquivos na pasta de entrada
    for filename in os.listdir(input_folder):
        if filename.endswith('.xlsx'):
            input_path = os.path.join(input_folder, filename)
            output_path = os.path.join(output_folder, filename)

            print(f"Processando: {filename}")

            # Carregue o workbook
            wb = load_workbook(input_path)

            # Percorra todas as planilhas no workbook
            for ws in wb.worksheets:
                # Aplique quebra de texto automática a todas as células
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True)

                # Ajuste a largura das colunas
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Salve o workbook modificado
            wb.save(output_path)
            print(f"Salvo como: {output_path}")

    print("Processamento concluído!")

# Uso do script
input_folder = r"C:\Users\VMContabil\Documents\Python automate\dadosExtract\Relatório chat onvio"
output_folder = r"C:\Users\VMContabil\Documents\Python automate\dadosExtract\Relatório chat onvio\Relatório chat onvio formatado"

apply_wrap_text(input_folder, output_folder)